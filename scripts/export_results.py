"""Writers for Markdown / CSV / xlsx outputs."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from src.core.schema import (
    ContentDraft,
    ScoredItem,
    SearchCitationResult,
    StructuredPost,
    TrendSummary,
    VideoPrompt,
)


# --------------------------------------------------------------------- markdown


def write_daily_report_md(
    *,
    out_path: Path,
    date_str: str,
    provider: str,
    llm_provider: str,
    top10: list[ScoredItem],
    trends: list[TrendSummary],
    fallback_used: list[str],
    warnings: list[str],
    run_id: str = "",
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append(f"# X Intelligence Daily Report — {date_str}\n")
    lines.append(f"- run_id: `{run_id}`" if run_id else "- run_id: `(unset)`")
    lines.append(f"- provider: `{provider}`")
    lines.append(f"- llm_provider: `{llm_provider}`")
    lines.append(
        f"- fallback_used: {', '.join(fallback_used) if fallback_used else '_(none — real LLM was reached)_'}"
    )
    if warnings:
        lines.append(f"- ⚠️ warnings: {len(warnings)} 件 (manifest.warnings 参照)")
    lines.append("\n## Top 10\n")
    for i, item in enumerate(top10, 1):
        lines.append(f"### {i}. score={item.score.total} ({item.score.method})")
        if isinstance(item.post, StructuredPost):
            who = item.post.author_handle or item.post.author or "(unknown)"
            lines.append(f"- {who} — {item.post.text}")
            lines.append(f"- URL: {item.post.url}")
        else:
            lines.append(f"- citation summary: {item.post.summary}")
            for u in item.post.cited_urls:
                lines.append(f"  - {u}")
        v = item.post.verification
        lines.append(
            f"- verification: `{v.verification_status}` / source: `{v.source_type}`"
            + (f" / risks: {', '.join(v.risk_flags)}" if v.risk_flags else "")
        )
        if item.why_important:
            lines.append(f"- なぜ重要: {item.why_important}")
        lines.append("")

    lines.append("\n## Trend Summary\n")
    for t in trends:
        lines.append(f"### {t.topic_label} ({t.item_count} items, sentiment={t.sentiment})")
        lines.append(t.short_jp_explanation)
        if t.emerging_keywords:
            lines.append(f"- emerging: {', '.join(t.emerging_keywords)}")
        if t.content_angles:
            lines.append("- content angles:")
            for a in t.content_angles:
                lines.append(f"  - {a}")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def write_drafts_md(out_dir: Path, drafts: list[ContentDraft]) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for i, d in enumerate(drafts, 1):
        path = out_dir / f"{i:02d}_{d.channel}.md"
        source_lines = [f"- {u}" for u in d.source_urls] if d.source_urls else ["- (none)"]
        lines = [
            f"# {d.channel} draft ({d.tone})",
            "",
            f"- needs_review: **{d.needs_review}**",
            f"- originality_note: {d.originality_note}",
            "",
            "## source_urls",
            *source_lines,
            "",
            "## source_summary",
            d.source_summary,
            "",
            "## my_angle",
            d.my_angle,
            "",
            "## draft_text",
            d.draft_text,
        ]
        if d.note_title_candidates:
            lines += ["", "## note_title_candidates", *[f"- {t}" for t in d.note_title_candidates]]
        if d.note_outline:
            lines += ["", "## note_outline", *[f"- {o}" for o in d.note_outline]]
        path.write_text("\n".join(lines), encoding="utf-8")
        written.append(path)
    return written


def write_video_prompts_md(out_dir: Path, prompts: list[VideoPrompt]) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    for i, p in enumerate(prompts, 1):
        path = out_dir / f"{i:02d}_{p.use_case}.md"
        lines = [
            f"# {p.use_case} — video/image prompt",
            "",
            "## Concept",
            p.concept,
            "",
            "## Grok Imagine Prompt (EN)",
            "```",
            p.grok_imagine_prompt_en,
            "```",
            "",
            "## 日本語説明",
            p.jp_explanation,
            "",
            "## Parameters",
            f"- duration: {p.duration}",
            f"- aspect_ratio: {p.aspect_ratio}",
            f"- mood: {p.mood}",
            f"- color_palette: {p.color_palette}",
            f"- camera: {p.camera_movement}",
            f"- text_overlay: {p.text_overlay}",
            f"- negative_prompt: {p.negative_prompt}",
        ]
        path.write_text("\n".join(lines), encoding="utf-8")
        written.append(path)
    return written


def write_review_queue(
    out_dir: Path,
    drafts: list[ContentDraft],
) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    for bucket in ("approved", "rejected", "needs_fact_check"):
        (out_dir / bucket).mkdir(exist_ok=True)

    queue_path = out_dir / "drafts_to_review.md"
    lines = ["# Drafts to Review", "", "レビュー後、該当ファイルを approved/ rejected/ needs_fact_check/ に手動移動してください。", ""]
    for i, d in enumerate(drafts, 1):
        lines.append(f"- [{i:02d}] {d.channel} (tone={d.tone}) — sources: {len(d.source_urls)}")
    queue_path.write_text("\n".join(lines), encoding="utf-8")
    return queue_path


# --------------------------------------------------------------------- csv


def write_csv(out_path: Path, items: Iterable[ScoredItem]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    rows = []
    for item in items:
        p = item.post
        if isinstance(p, StructuredPost):
            kind = "structured_post"
            text = p.text
            author = p.author_handle or p.author or ""
            created = p.created_at.isoformat() if p.created_at else ""
            url = p.url
            citations = ""
        else:
            kind = "search_citation_result"
            text = p.summary
            author = ""
            created = ""
            url = p.cited_urls[0] if p.cited_urls else ""
            citations = "|".join(p.cited_urls)
        v = p.verification
        rows.append({
            "kind": kind,
            "topic": p.topic or "",
            "score_total": item.score.total,
            "score_method": item.score.method,
            "importance": item.score.importance,
            "novelty": item.score.novelty,
            "career_relevance": item.score.career_relevance,
            "note_fit": item.score.note_fit,
            "x_fit": item.score.x_fit,
            "author": author,
            "created_at": created,
            "text": text,
            "url": url,
            "citations": citations,
            "verification_status": v.verification_status,
            "source_type": v.source_type,
            "risk_flags": "|".join(v.risk_flags),
            "missing_fields": "|".join(p.missing_fields),
            "why_important": item.why_important,
        })
    if not rows:
        rows = [{"kind": "", "topic": "", "score_total": "", "score_method": "", "importance": "",
                 "novelty": "", "career_relevance": "", "note_fit": "", "x_fit": "", "author": "",
                 "created_at": "", "text": "", "url": "", "citations": "", "verification_status": "",
                 "source_type": "", "risk_flags": "", "missing_fields": "", "why_important": ""}]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


# --------------------------------------------------------------------- xlsx


def write_xlsx(
    out_path: Path,
    *,
    top10: list[ScoredItem],
    all_items: list[ScoredItem],
    trends: list[TrendSummary],
    drafts: list[ContentDraft],
    video_prompts: list[VideoPrompt],
) -> None:
    from openpyxl import Workbook

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Top10
    ws = wb.active
    ws.title = "Top10"
    _scored_sheet(ws, top10)

    # AllItems
    ws2 = wb.create_sheet("AllItems")
    _scored_sheet(ws2, all_items)

    # TrendSummary
    ws3 = wb.create_sheet("TrendSummary")
    ws3.append(["topic_id", "topic_label", "item_count", "sentiment", "emerging_keywords", "short_jp", "content_angles"])
    for t in trends:
        ws3.append([
            t.topic_id, t.topic_label, t.item_count, t.sentiment,
            ", ".join(t.emerging_keywords), t.short_jp_explanation,
            "\n".join(t.content_angles),
        ])

    # ContentIdeas
    ws4 = wb.create_sheet("ContentIdeas")
    ws4.append(["channel", "tone", "needs_review", "source_urls", "source_summary", "my_angle", "draft_text", "originality_note"])
    for d in drafts:
        ws4.append([
            d.channel, d.tone, d.needs_review,
            "\n".join(d.source_urls), d.source_summary, d.my_angle,
            d.draft_text, d.originality_note,
        ])

    # VideoPrompts
    ws5 = wb.create_sheet("VideoPrompts")
    ws5.append(["use_case", "concept", "duration", "aspect_ratio", "mood", "color_palette", "grok_imagine_prompt_en", "jp_explanation"])
    for p in video_prompts:
        ws5.append([
            p.use_case, p.concept, p.duration, p.aspect_ratio,
            p.mood, p.color_palette, p.grok_imagine_prompt_en, p.jp_explanation,
        ])

    wb.save(out_path)


def _scored_sheet(ws, items: list[ScoredItem]) -> None:
    ws.append([
        "score_total", "score_method", "importance", "novelty", "career_relevance",
        "note_fit", "x_fit", "topic", "author", "created_at", "text",
        "url_or_first_citation", "all_citations",
        "verification_status", "source_type", "risk_flags", "missing_fields", "why_important",
    ])
    for item in items:
        p = item.post
        if isinstance(p, StructuredPost):
            author = p.author_handle or p.author or ""
            created = p.created_at.isoformat() if p.created_at else ""
            text = p.text
            url = p.url
            citations = p.url
        else:
            author = ""
            created = ""
            text = p.summary
            url = p.cited_urls[0] if p.cited_urls else ""
            citations = "\n".join(p.cited_urls)
        v = p.verification
        ws.append([
            item.score.total, item.score.method, item.score.importance, item.score.novelty,
            item.score.career_relevance, item.score.note_fit, item.score.x_fit,
            p.topic or "", author, created, text, url, citations,
            v.verification_status, v.source_type, ", ".join(v.risk_flags),
            ", ".join(p.missing_fields), item.why_important,
        ])
