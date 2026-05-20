from pathlib import Path

from scripts.export_results import write_csv, write_xlsx
from src.adapters.search_mock import MockSearchProvider
from src.core.dedupe import dedupe
from src.core.scoring import score_all, top_n
from src.utils.config_loader import load_config


def test_csv_written(tmp_path: Path):
    cfg = load_config()
    items = MockSearchProvider().search("any", "all", "24h").items
    scored = score_all(dedupe(items), profile=cfg.profile, weights=cfg.output["scoring"]["weights"])
    out = tmp_path / "out.csv"
    write_csv(out, scored)
    assert out.exists() and out.stat().st_size > 0
    content = out.read_text(encoding="utf-8").splitlines()
    assert content[0].startswith("kind,topic,score_total")


def test_xlsx_written_with_five_sheets(tmp_path: Path):
    from openpyxl import load_workbook
    cfg = load_config()
    items = MockSearchProvider().search("any", "all", "24h").items
    scored = score_all(dedupe(items), profile=cfg.profile, weights=cfg.output["scoring"]["weights"])
    top = top_n(scored, n=10)
    out = tmp_path / "report.xlsx"
    write_xlsx(out, top10=top, all_items=scored, trends=[], drafts=[], video_prompts=[])
    assert out.exists()
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {"Top10", "AllItems", "TrendSummary", "ContentIdeas", "VideoPrompts"}
